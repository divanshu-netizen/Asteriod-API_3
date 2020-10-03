import requests
import json
import jsonpath
import csv
import pytest
import openpyxl

# def test_default_query_parameters():
#     api_url = "https://ssd-api.jpl.nasa.gov/cad.api"
#     wk = openpyxl.load_workbook('C:/development/Git_Project/Asteriods_API/TestData/TestData.xlsx')
#     sh = wk['Sheet1']
#     rows = sh.max_row
#
#     for i in range(2,rows+1):
#         cell.date-min = sh.cell(row=i, column=1)
#         cell.date-max = sh.cell(row=i, column=2)
#         cell.dist-max = sh.cell(row=i, column=3)



def read_test_data_from_csv():
    test_data = []
    with open("C:\\development\\Git_Project\\Asteriods_API\\TestData\\data.csv", newline="") as csvfile:
        data = csv.reader(csvfile, delimiter=",")
        next(data) # skip header row
        for row in data:
            test_data.append(row)
        return test_data

x = read_test_data_from_csv()
print(x)

# @pytest.mark.parametrize("date-min, date-max, dist-min, dist-max", read_test_data_from_csv())
# def test_default_query_parameters(date-min, date-max, dist-min, dist-max):
#     response = requests.get(f"https://ssd-api.jpl.nasa.gov/cad.api/cad.api"/${date-min}/${date-max}/${dist-max})
#     response_body = response.json()
#     assert response_body["statuscode"][0]["status code"] == 200


